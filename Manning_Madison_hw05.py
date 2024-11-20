import openpyxl
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter

art_book = openpyxl.Workbook()
sheet = art_book.active

#Setting cell sizes
for row in range(1, 55):
    sheet.row_dimensions[row].height = 6
for column in range(1,53):
    column_letter = get_column_letter(column)
    sheet.column_dimensions[column_letter].width = 1
art_book.save("new_art.xlsx")

#Setting cell colors
black = 'FF010101'
fill_blk = PatternFill(patternType = 'solid', fgColor= black)
black_list = ['R1', 'S1', 'T1', 'U1', 'V1', 'Q2', 'W2', 'Q3', 'W3', 'X3', 'Q4', 'X4', 'Q5', 'X5', 'O6', 'P6', 'W6', 'X6', 
'M7', 'N7', 'Y7', 'Z7', 'K8', 'L8', 'AA8', 'AB8', 'J9', 'AB9', 'AC9', 'I10', 'AD10', 'H11', 'W11', 'X11', 'Y11', 'AE11', 
'H12', 'AF12', 'AG12', 'AH12', 'G13', 'AI13', 'AJ13', 'G14', 'AJ14', 'AK14', 'Z15', 'AA15', 'AK15', 'A16', 'Z16', 'AA16', 
'AB16', 'AK16', 'AL16', 'AM16', 'AN16', 'AO16', 'AP16', 'AQ16', 'A17', 'AA17', 'AB17', 'AH17', 'AI17', 'AJ17', 'AR17', 'AS17', 
'AT17', 'AU17', 'A18', 'AG18', 'AH18', 'AK18', 'AQ18', 'AV18', 'A19', 'AF19', 'AJ19', 'AR19', 'AV19', 'B20', 'F20', 'K20', 
'AF20', 'AK20', 'AL20', 'AM20', 'AN20', 'AO20', 'AP20', 'AQ20', 'AV20', 'C21', 'D21', 'E21', 'F21', 'K21', 'AF21', 'AG21', 
'AT21', 'AU21', 'F22', 'K22', 'Y22', 'AG22', 'AH22', 'AJ22', 'AM22', 'AN22', 'AQ22', 'AR22', 'AS22', 'AU22', 'G23', 'X23', 
'Y23', 'Z23', 'AG23', 'AI23', 'AK23', 'AM23', 'AN23', 'AQ23', 'AV23', 'G24', 'N24', 'O24', 'W24', 'X24', 'Y24', 'Z24', 'AF24', 
'AK24', 'AM24', 'AO24', 'AP24', 'AV24', 'H25', 'N25', 'O25', 'P25', 'X25', 'Y25', 'Z25', 'AD25', 'AF25', 'AK25', 'AM25', 'AV25', 
'H26', 'O26', 'P26', 'AD26', 'AF26', 'AK26', 'AL26', 'AW26', 'I27', 'AD27', 'AF27', 'AW27', 'AX27', 'J28', 'AC28', 'AF28', 'AH28', 
'AI28', 'AJ28', 'AK28', 'AL28', 'AW28', 'AX28', 'K29', 'Y29', 'Z29', 'AA29', 'AB29', 'AE29', 'AF29', 'AG29', 'AM29', 'AV29', 
'AY29', 'K30', 'AD30', 'AH30', 'AN30', 'AU30', 'AY30', 'K31', 'AB31', 'AC31', 'AH31', 'AN31', 'AT31', 'AY31', 'K32', 'L32', 
'AA32', 'AI32', 'AN32', 'AR32', 'AS32', 'AX32', 'L33', 'Z33', 'AI33', 'AM33', 'AN33', 'AO33', 'AP33', 'AQ33', 'AR33', 'AS33', 
'AW33', 'M34', 'Y34', 'AI34', 'AM34', 'AN34', 'AS34', 'AT34', 'AU34', 'AV34', 'N35', 'Y35', 'AJ35', 'AK35', 'AL35', 'AO35', 
'AP35', 'AQ35', 'AR35', 'O36', 'P36', 'W36', 'X36', 'Y36', 'AI36', 'AJ36', 'AP36', 'P37', 'Q37', 'R37', 'S37', 'T37', 'U37', 
'V37', 'Y37', 'AH37', 'AI37', 'AP37', 'V38', 'AE38', 'AF38', 'AG38', 'AM38', 'AN38', 'AO38', 'AP38', 'V39', 'AH39', 'AI39', 
'AJ39', 'AK39', 'AL39', 'AP39', 'V40', 'AE40', 'AF40', 'AG40', 'AP40', 'V41', 'AB41', 'AC41', 'AD41', 'AP41', 'V42', 'Z42', 
'AA42', 'AP42', 'V43', 'W43', 'X43', 'Y43', 'AP43', 'V44', 'AO44', 'V45', 'AO45', 'V46', 'AO46', 'W47', 'AD47', 'AE47', 'AF47', 
'AG47', 'AH47', 'AO47', 'W48', 'AD48', 'AH48', 'AN48', 'AO48', 'W49', 'AD49', 'AH49', 'AN49', 'W50', 'AD50', 'AG50', 'AN50', 
'X51', 'AE51', 'AG51', 'AN51', 'X52', 'AE52', 'AG52', 'AM52', 'Y53', 'Z53', 'AA53', 'AB53', 'AC53', 'AD53', 'AE53', 'AG53', 
'AH53', 'AI53', 'AJ53', 'AK53', 'AL53', 'AM53']

black_rgb = Color(rgb= '010101')

for cell in black_list:
    fill = PatternFill(patternType= 'solid', fgColor=black_rgb)
    art_book.save("new_art.xlsx")


yellow = 'FFFFCD2D'
fill_yel = PatternFill(patternType = 'solid', fgColor=yellow)
yellow_list = ['R2', 'S2', 'T2', 'U2', 'V2', 'R3', 'S3', 'T3', 'U3', 'V3', 'R4', 'S4', 'T4', 'U4', 'V4', 'W4', 'R5', 'S5', 
'T5', 'U5', 'V5', 'W5', 'Q6', 'R6', 'S6', 'T6', 'U6', 'V6', 'O7', 'P7', 'Q7', 'R7', 'S7', 'T7', 'U7', 'V7', 'W7', 'X7', 'M8', 
'N8', 'O8', 'P8', 'Q8', 'R8', 'S8', 'T8', 'U8', 'V8', 'W8', 'X8', 'Y8', 'Z8', 'K9', 'L9', 'M9', 'N9', 'O9', 'P9', 'Q9', 'R9', 
'S9', 'T9', 'U9', 'V9', 'W9', 'X9', 'Y9', 'Z9', 'AA9', 'J10', 'K10', 'L10', 'M10', 'N10', 'O10', 'P10', 'Q10', 'R10', 'S10', 
'T10', 'U10', 'V10', 'W10', 'X10', 'Y10', 'Z10', 'AA10', 'AB10', 'AC10', 'I11', 'J11', 'K11', 'L11', 'M11', 'N11', 'O11', 'P11', 
'Q11', 'R11', 'S11', 'T11', 'U11', 'V11', 'Z11', 'AA11', 'AB11', 'AC11', 'AD11', 'I12', 'J12', 'K12', 'L12', 'M12', 'N12', 'O12', 
'P12', 'Q12', 'R12', 'S12', 'T12', 'U12', 'V12', 'W12', 'X12', 'Y12', 'Z12', 'AA12', 'AB12', 'AC12', 'AD12', 'AE12', 'H13', 
'I13', 'J13', 'K13', 'L13', 'M13', 'N13', 'O13', 'P13', 'Q13', 'R13', 'S13', 'T13', 'U13', 'V13', 'W13', 'X13', 'Y13', 'Z13', 
'AA13', 'AB13', 'AC13', 'AD13', 'AE13', 'AF13', 'AG13', 'AH13', 'H14', 'I14', 'J14', 'K14', 'L14', 'M14', 'N14', 'O14', 'P14', 
'Q14', 'R14', 'S14', 'T14', 'U14', 'V14', 'W14', 'X14', 'Y14', 'Z14', 'AA14', 'AB14', 'AC14', 'AD14', 'AE14', 'AF14', 'AG14', 
'AH14', 'AI14', 'G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15', 'S15', 'T15', 'U15', 'V15', 
'W15', 'X15', 'Y15', 'AB15', 'AC15', 'AD15', 'AE15', 'AF15', 'AG15', 'AH15', 'AI15', 'AJ15', 'B16', 'C16', 'D16', 'E16', 'F16', 
'G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'P16', 'Q16', 'R16', 'S16', 'T16', 'U16', 'V16', 'W16', 'X16', 
'Y16', 'AC16', 'AD16', 'AE16', 'AF16', 'AG16', 'AH16', 'AI16', 'AJ16', 'B17', 'C17', 'D17', 'E17', 'F17', 'G17', 'H17', 'I17', 
'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'P17', 'Q17', 'R17', 'S17', 'T17', 'U17', 'V17', 'W17', 'X17', 'Y17', 'Z17', 'AC17', 
'AD17', 'AE17', 'AF17', 'AG17', 'B18', 'C18', 'D18', 'E18', 'F18', 'G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 
'P18', 'Q18', 'R18', 'S18', 'T18', 'U18', 'V18', 'W18', 'X18', 'Y18', 'Z18', 'AA18', 'AB18', 'AC18', 'AD18', 'AE18', 'AF18', 
'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'P19', 'Q19', 'R19', 'S19', 
'T19', 'U19', 'V19', 'W19', 'X19', 'Y19', 'Z19', 'AA19', 'AB19', 'AC19', 'AD19', 'AE19', 'C20', 'D20', 'E20', 'G20', 'H20', 
'I20', 'J20', 'L20', 'M20', 'N20', 'O20', 'P20', 'Q20', 'R20', 'S20', 'T20', 'U20', 'V20', 'W20', 'X20', 'Y20', 'Z20', 'AA20', 
'AB20', 'AC20', 'AD20', 'AE20', 'G21', 'H21', 'I21', 'J21', 'L21', 'M21', 'N21', 'O21', 'P21', 'Q21', 'R21', 'S21', 'T21', 'U21', 
'V21', 'W21', 'X21', 'Y21', 'Z21', 'AA21', 'AB21', 'AC21', 'AD21', 'AE21', 'G22', 'H22', 'I22', 'J22', 'L22', 'M22', 'N22', 
'O22', 'P22', 'Q22', 'R22', 'S22', 'T22', 'U22', 'V22', 'W22', 'X22', 'Z22', 'AA22', 'AB22', 'AC22', 'AD22', 'AE22', 'AF22', 
'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'Q23', 'R23', 'S23', 'T23', 'U23', 'V23', 'W23', 'AA23', 'AB23', 
'AC23', 'AD23', 'AE23', 'AF23', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'P24', 'Q24', 'R24', 'S24', 'T24', 'U24', 'V24', 
'AA24', 'AB24', 'AC24', 'AD24', 'AE24', 'I25', 'J25', 'K25', 'L25', 'M25', 'Q25', 'R25', 'S25', 'T25', 'U25', 'V25', 'W25', 
'AA25', 'AB25', 'AC25', 'AE25', 'I26', 'J26', 'K26', 'L26', 'M26', 'N26', 'Q26', 'R26', 'S26', 'T26', 'U26', 'V26', 'W26', 
'X26', 'Y26', 'Z26', 'AA26', 'AB26', 'AC26', 'AE26', 'J27', 'K27', 'L27', 'M27', 'N27', 'O27', 'P27', 'Q27', 'R27', 'S27', 
'T27', 'U27', 'V27', 'W27', 'X27', 'Y27', 'Z27', 'AA27', 'AB27', 'AC27', 'AE27', 'K28', 'L28', 'M28', 'N28', 'O28', 'P28', 
'Q28', 'R28', 'S28', 'T28', 'U28', 'V28', 'W28', 'X28', 'Y28', 'Z28', 'AA28', 'AB28', 'AD28', 'AE28', 'L29', 'M29', 'N29', 'O29', 
'P29', 'Q29', 'R29', 'S29', 'T29', 'U29', 'V29', 'W29', 'X29', 'AC29', 'AD29', 'AH29', 'AI29', 'AJ29', 'AK29', 'AL29', 'AW29', 
'AX29', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30', 'S30', 'T30', 'U30', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 
'AC30', 'AI30', 'AJ30', 'AK30', 'AL30', 'AM30', 'AV30', 'AW30', 'AX30', 'L31', 'M31', 'N31', 'O31', 'P31', 'Q31', 'R31', 'S31', 
'T31', 'U31', 'V31', 'W31', 'X31', 'Y31', 'Z31', 'AA31', 'AI31', 'AJ31', 'AK31', 'AL31', 'AM31', 'AU31', 'AV31', 'AW31', 'AX31', 
'M32', 'N32', 'O32', 'P32', 'Q32', 'R32', 'S32', 'T32', 'U32', 'V32', 'W32', 'X32', 'Y32', 'Z32', 'AJ32', 'AK32', 'AL32', 'AM32', 
'AT32', 'AU32', 'AV32', 'AW32', 'M33', 'N33', 'O33', 'P33', 'Q33', 'R33', 'S33', 'T33', 'U33', 'V33', 'W33', 'X33', 'Y33', 
'AJ33', 'AK33', 'AL33', 'AT33', 'AU33', 'AV33', 'N34', 'O34', 'P34', 'Q34', 'R34', 'S34', 'T34', 'U34', 'V34', 'W34', 'X34', 
'AJ34', 'AK34', 'AL34', 'O35', 'P35', 'Q35', 'R35', 'S35', 'T35', 'U35', 'V35', 'W35', 'X35', 'Q36', 'R36', 'S36', 'T36', 'U36', 
'V36', 'AM39', 'AN39', 'AO39', 'AH40', 'AI40', 'AJ40', 'AK40', 'AL40', 'AM40', 'AN40', 'AO40', 'AE41', 'AF41', 'AG41', 'AH41', 
'AI41', 'AJ41', 'AK41', 'AL41', 'AM41', 'AN41', 'AO41', 'AB42', 'AC42', 'AD42', 'AE42', 'AF42', 'AG42', 'AH42', 'AI42', 'AJ42', 
'AK42', 'AL42', 'AM42', 'AN42', 'AO42', 'Z43', 'AA43', 'AB43', 'AC43', 'AD43', 'AE43', 'AF43', 'AG43', 'AH43', 'AI43', 'AJ43', 
'AK43', 'AL43', 'AM43', 'AN43', 'AO43', 'W44', 'X44', 'Y44', 'Z44', 'AA44', 'AB44', 'AC44', 'AD44', 'AE44', 'AF44', 'AG44', 
'AH44', 'AI44', 'AJ44', 'AK44', 'AL44', 'AM44', 'AN44', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45', 
'AF45', 'AG45', 'AH45', 'AI45', 'AJ45', 'AK45', 'AL45', 'AM45', 'AN45', 'W46', 'X46', 'Y46', 'Z46', 'AA46', 'AB46', 'AC46', 
'AD46', 'AE46', 'AF46', 'AG46', 'AH46', 'AI46', 'AJ46', 'AK46', 'AL46', 'AM46', 'AN46', 'X47', 'Y47', 'Z47', 'AA47', 'AB47', 
'AC47', 'AI47', 'AJ47', 'AK47', 'AL47', 'AM47', 'AN47', 'X48', 'Y48', 'Z48', 'AA48', 'AB48', 'AC48', 'AI48', 'AJ48', 'AK48', 
'AL48', 'AM48', 'X49', 'Y49', 'Z49', 'AA49', 'AB49', 'AC49', 'AI49', 'AJ49', 'AK49', 'AL49', 'AM49', 'X50', 'Y50', 'Z50', 'AA50', 
'AB50', 'AC50', 'AH50', 'AI50', 'AJ50', 'AK50', 'AL50', 'AM50', 'Y51', 'Z51', 'AA51', 'AB51', 'AC51', 'AD51', 'AH51', 'AI51', 
'AJ51', 'AK51', 'AL51', 'AM51', 'Y52', 'Z52', 'AA52', 'AB52', 'AC52', 'AD52', 'AH52', 'AI52', 'AJ52', 'AK52', 'AL52']

yellow_rgb = Color(rgb= 'FFCD2D')

for cell in yellow_list:
    fill = PatternFill(patternType= 'solid', fgColor=yellow_rgb)
    art_book.save("new_art.xlsx")

pale_yellow = 'FFFFE285'
fill_pyel = PatternFill(patternType = 'solid', fgColor=pale_yellow)
pale_yellow_list = ['AK17', 'AL17', 'AM17', 'AN17', 'AO17', 'AP17', 'AQ17', 'AI18', 'AJ18', 'AL18', 'AM18', 'AN18', 'AO18', 
'AP18', 'AR18', 'AS18', 'AT18', 'AU18', 'AG19', 'AH19', 'AI19', 'AK19', 'AL19', 'AM19', 'AN19', 'AO19', 'AP19', 'AQ19', 'AS19', 
'AT19', 'AU19', 'AG20', 'AH20', 'AI20', 'AJ20', 'AR20', 'AS20', 'AT20', 'AU20', 'AH21', 'AI21', 'AJ21', 'AK21', 'AL21', 'AM21', 
'AN21', 'AO21', 'AP21', 'AQ21', 'AR21', 'AS21', 'AI22', 'AK22', 'AL22', 'AO22', 'AP22', 'AL23', 'AO23', 'AP23', 'AL24', 'AL25']

pale_yellow_rgb = Color(rgb= 'FFE285')

for cell in pale_yellow_list:
    fill = PatternFill(patternType= 'solid', fgColor=pale_yellow_rgb)
    art_book.save("new_art.xlsx")

red = 'FFC92403'
fill_red = PatternFill(patternType = 'solid', fgColor=red)
red_list = ['AE30', 'AF30', 'AG30', 'AD31', 'AE31', 'AF31', 'AG31', 'AB32', 'AC32', 'AD32', 'AE32', 'AF32', 'AG32', 'AH32', 
'AA33', 'AB33', 'AC33', 'AD33', 'AE33', 'AF33', 'AG33', 'AH33', 'Z34', 'AA34', 'AB34', 'AC34', 'AD34', 'AE34', 'AF34', 'AG34', 
'AH34', 'AO34', 'AP34', 'AQ34', 'AR34', 'Z35', 'AA35', 'AB35', 'AC35', 'AD35', 'AE35', 'AF35', 'AG35', 'AH35', 'AI35', 'AM35', 
'AN35', 'Z36', 'AA36', 'AB36', 'AC36', 'AD36', 'AE36', 'AF36', 'AG36', 'AH36', 'AK36', 'AL36', 'AM36', 'AN36', 'AO36', 'W37', 
'X37', 'Z37', 'AA37', 'AB37', 'AC37', 'AD37', 'AE37', 'AF37', 'AG37', 'AJ37', 'AK37', 'AL37', 'AM37', 'AN37', 'AO37', 'W38', 
'X38', 'Y38', 'Z38', 'AA38', 'AB38', 'AC38', 'AD38', 'AH38', 'AI38', 'AJ38', 'AK38', 'AL38', 'W39', 'X39', 'Y39', 'Z39', 'AA39', 
'AB39', 'AC39', 'AD39', 'AE39', 'AF39', 'AG39', 'W40', 'X40', 'Y40', 'Z40', 'AA40', 'AB40', 'AC40', 'AD40', 'W41', 'X41', 'Y41', 
'Z41', 'AA41', 'W42', 'X42', 'Y42']

red_rgb = Color(rgb= 'C92403')

for cell in red_list:
    fill = PatternFill(patternType= 'solid', fgColor=red_rgb)
    art_book.save("new_art.xlsx")

blue = 'FF99ACD1'
fill_bl = PatternFill(patternType = 'solid', fgColor=blue)
blue_list = ['AT22', 'AH23', 'AJ23', 'AR23', 'AS23', 'AT23', 'AU23', 'AG24', 'AH24', 'AI24', 'AJ24', 'AN24', 'AQ24', 'AR24', 
'AS24', 'AT24', 'AU24', 'AG25', 'AH25', 'AI25', 'AJ25', 'AN25', 'AO25', 'AP25', 'AQ25', 'AR25', 'AS25', 'AT25', 'AU25', 'AG26', 
'AH26', 'AI26', 'AJ26', 'AM26', 'AN26', 'AO26', 'AP26', 'AQ26', 'AR26', 'AS26', 'AT26', 'AU26', 'AV26', 'AG27', 'AH27', 'AI27', 
'AJ27', 'AK27', 'AL27', 'AM27', 'AN27', 'AO27', 'AP27', 'AQ27', 'AR27', 'AS27', 'AT27', 'AU27', 'AV27', 'AG28', 'AM28', 'AN28', 
'AO28', 'AP28', 'AQ28', 'AR28', 'AS28', 'AT28', 'AU28', 'AV28', 'AN29', 'AO29', 'AP29', 'AQ29', 'AR29', 'AS29', 'AT29', 'AU29', 
'AO30', 'AP30', 'AQ30', 'AR30', 'AS30', 'AT30', 'AO31', 'AP31', 'AQ31', 'AR31', 'AS31', 'AO32', 'AP32', 'AQ32']

blue_rgb = Color(rgb= '99ACD1')

for cell in blue_list:
    fill = PatternFill(patternType= 'solid', fgColor=blue_rgb)
    art_book.save("new_art.xlsx")

#scraper
'''

wb = openpyxl.load_workbook("COSC1010_pixel_art.xlsx")
sheet = wb.active
cells = sheet['A1:BA55']
color_dic = {'blk': [], 'yel' : [], 'p_yel' : [], 'r' : [], 'bl' : []}

for row in sheet['A1:BA55']:
    for cell in row:
        if cell.fill and cell.fill.fgColor:
            color = cell.fill.fgColor.rgb
            print(f"color: {color}")
            if color == black:
                color_dic['blk'].append(cell.coordinate)
            elif color == yellow:
                color_dic['yel'].append(cell.coordinate)
            elif color == pale_yellow:
                color_dic['p_yel'].append(cell.coordinate)
            elif color == red:
                color_dic['r'].append(cell.coordinate)
            elif color == blue:
                color_dic['bl'].append(cell.coordinate)


print(f"Yellow:", color_dic['bl'])
        
'''
